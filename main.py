import webview
from flask import Flask, render_template
from markupsafe import Markup, escape
import pysrt
import xlsxwriter
from openpyxl import load_workbook

import os
import time
from math import ceil, floor
from AppKit import NSScreen


screenW = NSScreen.mainScreen().frame().size.width
screenH = NSScreen.mainScreen().frame().size.height
appScreenRatio = 0.5
msg = ""
msg += Markup("Many SRT Files can convert into one Excel File with worksheets.<br>")
msg += Markup("One Excel File with many worksheets can convert into many SRT Files.<br>")

app = Flask(__name__, static_folder='assets')

@app.route('/')
def main():
    return render_template('index.html', Message=msg)

class Api:
    def __init__(self):
        self._window = None

    def set_window(self, window):
        self._window = window

    def quit(self):
        self._window.destroy()

    def minimize(self):
        self._window.minimize()

    def enlarge(self):
        self._window.move(0, 0)
        self._window.resize(screenW, screenH)

    def import_srt(self):
        file_types = ('SRT Files (*.srt)', 'All files (*.*)')
        filelist = self._window.create_file_dialog(webview.OPEN_DIALOG, allow_multiple=True, file_types=file_types)
        read_srt(filelist)

    def import_excel(self):
        file_types = ('Excel File (*.xlsx;*.xls)', 'All files (*.*)')
        file = self._window.create_file_dialog(webview.OPEN_DIALOG, allow_multiple=False, file_types=file_types)
        read_excel(file)

def start_app():
    api = Api()
    window = webview.create_window(
        title='Film Assist',
        url=app,
        js_api=api,
        width=screenW*appScreenRatio, height=screenH*appScreenRatio, min_size=(320, 480),
        x=None, y=None, hidden=False,
        fullscreen=False, resizable=True,
        frameless=True, easy_drag=False, text_select=False,
        minimized=False, on_top=False,
        confirm_close=True, background_color='#2E2E2E'
    )
    api.set_window(window)
    webview.start()

def read_srt(seqList):
    if len(seqList) > 0:
        seqid = 1
        srt_pathname_first, srt_pathext_first = os.path.splitext(seqList[0])
        seqNameFirst = os.path.basename(srt_pathname_first).split('.')[0]
        workbook_path = srt_pathname_first + '_FilmAssist' + '.xlsx'
        workbook_filename_e = seqNameFirst + '_FilmAssist' + '.xlsx'
        workbook = xlsxwriter.Workbook(workbook_path)
        worksheet = {}
        for srt_path in seqList:
            srt_pathname, srt_pathext = os.path.splitext(srt_path)
            seqName = os.path.basename(srt_pathname).split('.')[0]
            # print('*'*80)
            # print('*'*2, srt_path, '*'*2)
            # print('*'*80)
            subs = pysrt.open(srt_path)
            if len(subs) != 0:
                # print("Test 1st row")
                # print(subs[0].start, subs[0].end, subs[0].text)
                # print('*'*80)

                worksheet[seqid] = workbook.add_worksheet(seqName)

                row = 0
                col = 0
                font_size = 16
                format_title = workbook.add_format({'font_name':'Arial', 'font_size':font_size, 'align':'center', 'valign':'vcenter','text_wrap':True, 'bold': True})
                format_content = workbook.add_format({'font_name':'Arial', 'font_size':font_size, 'align':'center', 'valign':'vcenter','text_wrap':True})
                worksheet[seqid].set_column('A:A', 20)
                worksheet[seqid].set_column('B:B', 20)
                worksheet[seqid].set_column('C:C', 80)

                format_current = format_title
                worksheet[seqid].write(row, col, 'TC Start', format_current)
                worksheet[seqid].write(row, col+1, 'TC End', format_current)
                worksheet[seqid].write(row, col+2, 'Subtitle', format_current)
                row += 1
                format_current = format_content

                for sub in subs:
                    worksheet[seqid].write(row, col, "{}".format(sub.start), format_current)
                    worksheet[seqid].write(row, col+1, "{}".format(sub.end), format_current)
                    worksheet[seqid].write(row, col+2, sub.text, format_current)
                    # print(sub.start, sub.end, sub.text)
                    row += 1

        workbook.close()

def read_excel(excels):
    excel = excels[0]
    wb_pathname, wb_pathext = os.path.splitext(excel)
    wb_dirname = os.path.dirname(excel)
    wb_filename = os.path.basename(wb_pathname).split('.')[0]
    # print(excel)
    wb = load_workbook(filename = excel)
    for ws in wb.worksheets:
        # print(ws.title)
        srt = pysrt.SubRipFile()
        id = 0
        for row in ws.rows:
            if id > 0 :
                sub = pysrt.SubRipItem(id, start=row[0].value, end=row[1].value, text=row[2].value)
                srt.append(sub)
            id += 1
        srt.save('{}/{}_FilmAssist.srt'.format(wb_dirname, ws.title))

if __name__ == '__main__':
    start_app()
